from __future__ import annotations

from typing import Any, List, Union

import openpyxl
import polars as pl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class XlsxHandler:
    @staticmethod
    def read(
        path: str,
        sheet_name: Union[str, int] = 0,
        has_header: bool = True,
        skip_rows: int = 0,
    ) -> pl.DataFrame:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if isinstance(sheet_name, int):
            names = wb.sheetnames
            if sheet_name >= len(names):
                wb.close()
                raise ValueError(f"Sheet index {sheet_name} out of range (sheets: {names})")
            sheet_name = names[sheet_name]
        ws = wb[sheet_name]

        rows: List[List[Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < skip_rows:
                continue
            rows.append(list(row))

        wb.close()

        if not rows:
            return pl.DataFrame()

        if has_header:
            header = [str(c) if c is not None else f"column_{i}" for i, c in enumerate(rows[0])]
            data = rows[1:]
        else:
            header = [f"column_{i}" for i in range(len(rows[0]))]
            data = rows

        max_cols = max(len(r) for r in data) if data else 0
        while len(header) < max_cols:
            header.append(f"column_{len(header)}")

        padded: List[List[Any]] = []
        for r in data:
            while len(r) < len(header):
                r.append(None)
            padded.append(r[: len(header)])

        return pl.DataFrame(padded, schema=header, orient="row")

    @staticmethod
    def list_sheets(path: str) -> List[str]:
        wb = openpyxl.load_workbook(path, read_only=True)
        names: List[str] = list(wb.sheetnames)
        wb.close()
        return names

    @staticmethod
    def write(
        df: pl.DataFrame,
        path: str,
        sheet_name: str = "Sheet1",
        include_header: bool = True,
        apply_table_style: bool = True,
    ) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = sheet_name

        header_fill = PatternFill(start_color="1e1e2e", end_color="1e1e2e", fill_type="solid")
        header_font = Font(bold=True, color="cdd6f4", size=11)
        data_font = Font(color="cdd6f4", size=10)
        thin_border = Border(
            left=Side(style="thin", color="313244"),
            right=Side(style="thin", color="313244"),
            top=Side(style="thin", color="313244"),
            bottom=Side(style="thin", color="313244"),
        )

        if include_header:
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            start_row = 2
        else:
            start_row = 1

        for row_idx, row in enumerate(df.iter_rows(), start_row):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="right" if isinstance(val, (int, float)) else "left"
                )

        for col_idx in range(1, len(df.columns) + 1):
            max_len = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                val = row[0]
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        if apply_table_style:
            from openpyxl.worksheet.table import Table, TableStyleInfo

            if ws.max_row > 1:
                ref = f"A1:{get_column_letter(len(df.columns))}{ws.max_row}"
                table = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}", ref=ref)
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                table.tableStyleInfo = style
                ws.add_table(table)

        wb.save(path)
        wb.close()
